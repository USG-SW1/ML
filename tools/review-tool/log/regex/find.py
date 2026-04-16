
import streamlit as st
import pandas as pd
import glob
import os
from io import BytesIO

st.title("Multi Regex Code Exact Match Exporter")


# Fixed CSV directory
csv_dir = "/home/gsbu/ml/tools/review-tool/log"
if not os.path.exists(csv_dir):
    st.warning(f"Directory does not exist: {csv_dir}")
csv_files = glob.glob(os.path.join(csv_dir, "*.csv"))

# Session state to store all matches and regex codes
if 'all_matches' not in st.session_state:
    st.session_state['all_matches'] = []
if 'regex_codes' not in st.session_state:
    st.session_state['regex_codes'] = set()


# User input for multiple regex codes
regex_codes_input = st.text_area("Paste regex codes (one per line):", "")
regex_codes = [line.strip() for line in regex_codes_input.strip().splitlines() if line.strip()]

if st.button("Find Matches") and regex_codes:
    st.session_state['all_matches'] = []
    st.session_state['regex_codes'] = set(regex_codes)
    for regex_code in regex_codes:
        for csv_path in csv_files:
            try:
                df = pd.read_csv(csv_path, encoding='utf-8-sig')
                if 'feature_raw' in df.columns:
                    first_line = df['feature_raw'].astype(str).str.split('\n').str[0]
                    mask = first_line == regex_code
                    matches = df[mask].copy()
                    if not matches.empty:
                        matches['regex_code'] = regex_code
                        st.session_state['all_matches'].append(matches)
            except Exception as e:
                st.write(f"Error reading {csv_path}: {e}")


if st.session_state['all_matches']:
    df_matches = pd.concat(st.session_state['all_matches'], ignore_index=True)
    st.write(f"Total matching rows for all regex codes: {len(df_matches)}")
    st.dataframe(df_matches)

    # Dashboard-style unique point analysis for each type
    st.subheader("Unique Points Dashboard")
    import random
    unique_codes = list(df_matches['regex_code'].unique())
    samples = {}
    for code in unique_codes:
        group = df_matches[df_matches['regex_code'] == code]
        if not group.empty:
            samples[code] = group.sample(1, random_state=42).iloc[0]

    # Find all unique points for each code
    unique_points_dict = {}
    for code, sample in samples.items():
        other_samples = [v for k, v in samples.items() if k != code]
        unique_points = {}
        for col in df_matches.columns:
            if col == 'regex_code':
                continue
            val = sample[col]
            if all(val != other[col] for other in other_samples):
                unique_points[col] = val
        unique_points_dict[code] = unique_points

    # Build dashboard table: rows=regex_code, columns=unique fields, cell=unique value
    all_unique_fields = set()
    for up in unique_points_dict.values():
        all_unique_fields.update(up.keys())
    all_unique_fields = sorted(all_unique_fields)
    dashboard_data = []
    for code in unique_codes:
        row = {'regex_code': code}
        for field in all_unique_fields:
            row[field] = unique_points_dict[code].get(field, "")
        dashboard_data.append(row)
    if dashboard_data and all_unique_fields:
        dashboard_df = pd.DataFrame(dashboard_data)
        st.dataframe(dashboard_df)
    else:
        st.write("No unique points found for any type compared to others.")
else:
    st.write("No matches found yet. Paste regex codes and click 'Find Matches'.")

import openpyxl
from openpyxl.styles import PatternFill

if st.button("Export All Matches to Excel"):
    if st.session_state['all_matches']:
        output = BytesIO()
        df_matches = pd.concat(st.session_state['all_matches'], ignore_index=True)
        # Write to Excel with color by regex_code
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_matches.to_excel(writer, index=False, sheet_name='Matches')
            wb = writer.book
            ws = writer.sheets['Matches']
            # Find the column index for 'regex_code'
            header = [cell.value for cell in ws[1]]
            try:
                regex_col_idx = header.index('regex_code') + 1  # openpyxl is 1-based
            except ValueError:
                regex_col_idx = ws.max_column  # fallback: last column
            # Assign a color for each regex_code
            unique_codes = list(df_matches['regex_code'].unique())
            import random
            def random_color():
                return ''.join([random.choice('0123456789ABCDEF') for _ in range(6)])
            # Start with a large palette, then add random colors if needed
            base_colors = [
                "FFFF99", "CCFFCC", "FFCCCC", "CCE5FF", "FFCCFF", "E0E0E0", "FFD580", "B3E6FF", "FFB3B3", "B3FFB3",
                "D9EAD3", "F4CCCC", "CFE2F3", "FCE5CD", "D9D2E9", "FFF2CC", "C9DAF8", "F9CB9C", "B6D7A8", "A2C4C9",
                "B4A7D6", "D5A6BD", "A4C2F4", "B7B7B7", "FFD966", "B6D7A8", "A2C4C9", "B4A7D6", "D5A6BD", "A4C2F4"
            ]
            colors = base_colors.copy()
            while len(colors) < len(unique_codes):
                # Generate a new random color not already in the list
                c = random_color()
                while c in colors:
                    c = random_color()
                colors.append(c)
            code_to_fill = {code: PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid") for i, code in enumerate(unique_codes)}
            # Color rows by regex_code (skip header row)
            for row in range(2, ws.max_row + 1):
                code = ws.cell(row=row, column=regex_col_idx).value
                if code in code_to_fill:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = code_to_fill[code]
        output.seek(0)
        st.download_button(
            label="Download Excel",
            data=output,
            file_name="all_regex_matches.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("No data to export. Please add regex codes and find matches first.")
